from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from datetime import datetime
import time as t
import openpyxl
import comtypes.client
import os
import PyPDF2
import xlwings as xw
from openpyxl import Workbook
from openpyxl.drawing.image import Image

class App:
    def __init__(self):
        self.app = Tk()
        self.app.title('檢貨單')
        self.frame1 = Frame(self.app)
        self.frame2 = Frame(self.app)
        self.frame3 = Frame(self.app)
        self.ent1_1 = None
        self.ent1_2 = None
        self.ent2_1 = None
        self.ent3_1 = None
        self.scrollbar_h = None
        self.selected_file_label = None
        self.excel_data = []
        self.current_path = ""
        self.progress = ttk.Progressbar(self.app, orient=HORIZONTAL, length=200, mode='determinate')
        self.progress.pack()
    def date(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path:
            # 显示所选文件的名称在Label上
            self.selected_file_label.config(text="選擇的檔案：" + file_path)
            self.read_excel(file_path)
    def read_excel(self, file_path):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            self.excel_data = []
            for row in sheet.iter_rows(values_only=True):
                self.excel_data.append(list(row))
            print(self.excel_data[1][2])
            workbook.close()

        except Exception as e:
            Label(self.frame1, text="讀取失敗,請再讀一次", font=('正楷', 18)).grid(column=0,row=0)
            print("讀取失敗,請再讀一次", str(e))




    def output(self):
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = f"檢貨單_{current_time}"
        folder_path = os.path.join(os.getcwd(), folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        self.current_path = folder_path
        file_path2 = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
        if file_path2:
            self.selected_file_label.config(text="空白檢貨單：" + file_path2)
        print(self.excel_data[1][2])
        workbook = openpyxl.load_workbook(file_path2)
        sheet = workbook.active

        print(len(self.excel_data))

        for row in range(len(self.excel_data)):
            img = Image('qrcode.png')
            if self.excel_data[row][0] == '回饋 ID' : continue
            sheet['B4'] = self.excel_data[row][2]
            sheet['B5'] = self.excel_data[row][7]
            sheet['B6'] = self.excel_data[row][8]
            sheet['B7'] = self.excel_data[row][14]
            sheet['B8'] = self.excel_data[row][9]
            sheet['G4'] = self.excel_data[row][16]
            sheet['E8'] = self.excel_data[row][10]
            sheet['G8'] = self.excel_data[row][11]
            sheet['A11'] = self.excel_data[row][0]
            sheet['B11'] = self.excel_data[row][1]
            sheet['G11'] = self.excel_data[row][15]
            sheet['B16'] = self.excel_data[row][19]
            sheet['E11'] = self.excel_data[row][20]
            sheet.add_image(img,'F20')
            filename = "檢貨單" + str(row) + ".xlsx"
            file_save_path = os.path.join(folder_path, filename)
            workbook.save(file_save_path)

            self.progress['value'] = (row + 1) / len(self.excel_data) * 100
            self.app.update_idletasks()
            workbook.close()
            #self.save_as_pdf(filename)
        # 保存工作簿到一個新的Excel文件
        messagebox.showinfo('完成','檢貨單已輸出完成!')
        # Excel 檔案的路徑列表
        #excel_files = ['檢貨單0.xlsx', '檢貨單1.xlsx', '檢貨單2.xlsx']

        # 讀取每個檔案並將它們存儲在 DataFrame 列表中
        #frames = [pd.read_excel(file) for file in excel_files]

        # 合併所有 DataFrame
        #merged_df = pd.concat(frames)

        # 將合併後的 DataFrame 保存為新的 Excel 檔案
        #merged_df.to_excel('merged.xlsx', index=False)
        print('檢貨單已保存')

        self.progress['value'] = 0


    def save_as_pdf(self):
        # 加載Excel

        excel = comtypes.client.CreateObject('Excel.Application')

        # 後台運行，不顯示
        excel.Visible = False

        # 禁止顯示彈窗
        excel.DisplayAlerts = False
        print("Current Working Directory:", os.getcwd())
        # 打開工作簿
        for i in range (len(self.excel_data)):
            if self.excel_data[i][0] == '回饋 ID': continue
            file_path = self.current_path + '\\檢貨單' + str(i) + '.xlsx'
            workbook = excel.Workbooks.Open(file_path)

            # 將工作簿保存為PDF
            pdf_path = file_path.replace('.xlsx', '.pdf')
            workbook.SaveAs(pdf_path, FileFormat=57)
            self.progress['value'] = (i + 1) / len(self.excel_data) * 33
            self.app.update_idletasks()
            # 關閉
            workbook.Close()

        pdf_writer = PyPDF2.PdfWriter()
        for j in range(len(self.excel_data)):
            if self.excel_data[j][0] == '回饋 ID': continue
            pdf_reader = PyPDF2.PdfReader(self.current_path + '\\檢貨單' + str(j) + '.pdf')
            pdf_writer.add_page(pdf_reader.pages[0])
            self.progress['value'] = (j + 1) / len(self.excel_data) * 33 + 33
            self.app.update_idletasks()
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        merge_name = "檢貨單" + current_time + ".pdf"
        merge_pdf_path = os.path.join(self.current_path,merge_name)
        with open(merge_pdf_path, 'wb') as out:
            pdf_writer.write(out)

        for x in range(len(self.excel_data)):
            if self.excel_data[x][0] == '回饋 ID': continue
            os.remove(self.current_path + '\\檢貨單' + str(x) + '.pdf')
            self.progress['value'] = (x + 1) / len(self.excel_data) * 33 + 67
            self.app.update_idletasks()
        # 退出Excel
        excel.Quit()

        messagebox.showinfo('完成', '檢貨單已合併完成!')
    def screen(self):
        sw = self.app.winfo_screenwidth()  # 得到屏幕宽度
        sh = self.app.winfo_screenheight()  # 得到屏幕高度

        ww, wh = 1000, 500
        x, y = (sw - ww) / 2, (sh - wh) / 2
        self.app.geometry(f"%dx%d+%d+%d" % (ww - 180, wh, x + 100, y))


    def buttons(self):
        Label(self.frame1, text="請匯入excel檔,匯出前需選擇空白檢貨單的excel檔", font=('正楷', 18)).grid(column=0, row=0)
        bt = Button(self.frame2, text='匯入', font=('正楷', 15), command=self.date, activebackground='blue',
                    overrelief='sunken')
        bt.grid(column=1, row=0)
        bt = Button(self.frame2, text='匯出一張張檢貨單', font=('正楷', 15), command=self.output, activebackground='blue',
                    overrelief='sunken')
        bt.grid(column=2, row=0)
        bt = Button(self.frame2, text='合成一份pdf', font=('正楷', 15), command=self.save_as_pdf, activebackground='blue',
                    overrelief='sunken')
        bt.grid(column=3, row=0)

        self.selected_file_label = Label(self.frame3, text="", font=('正楷', 10))
        self.selected_file_label.grid(row=1)
    def end(self):
        self.frame1.pack(anchor='w')
        self.frame2.pack(anchor='w')
        self.frame3.pack()
        self.app.resizable(width=False, height=False)  # 可不可拉伸

    def start(self):
        self.screen()
        #self.labels()
        #self.entry()
        self.buttons()
        self.end()
        self.app.mainloop()


if __name__ == '__main__':
    st = App()
    st.start()
    t.sleep(1)